"""Focused extraction from the annual NJ DCA summary sheets."""

from __future__ import annotations

import re
from collections.abc import Mapping
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import SOURCE_WORKBOOK_NAME, TARGET_MUNICIPALITIES, TARGET_YEARS

HEADER_ROW = 5
FIRST_DATA_ROW = 6


class ExtractionError(RuntimeError):
    """Raised when the workbook layout violates the expected source contract."""


def canonical_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        raise ExtractionError(f"Boolean municipality code is invalid: {value!r}")
    if isinstance(value, (int, float)):
        if int(value) != value:
            raise ExtractionError(f"Non-integral municipality code: {value!r}")
        return f"{int(value):04d}"
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.zfill(4) if text.isdigit() else text


def _header_map(values: tuple[Any, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        if value not in (None, ""):
            result[str(value).strip()] = index
    return result


def _require_header(headers: dict[str, int], label: str, sheet_name: str) -> int:
    try:
        return headers[label]
    except KeyError as exc:
        raise ExtractionError(
            f"{sheet_name} is missing required header {label!r} in row {HEADER_ROW}."
        ) from exc


def _require_matching_header(
    headers: dict[str, int],
    *,
    sheet_name: str,
    description: str,
    predicate: Any,
) -> tuple[str, int]:
    matches = [
        (label, column)
        for label, column in headers.items()
        if predicate(label)
    ]
    if len(matches) != 1:
        raise ExtractionError(
            f"{sheet_name} must contain exactly one {description} header in "
            f"row {HEADER_ROW}; found {matches}."
        )
    return matches[0]


def _cell_address(column: int, row: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _tax_reference_year(label: str, sheet_name: str) -> int:
    match = re.search(r"\bCY\s+(\d{4})\b", label)
    if not match:
        raise ExtractionError(
            f"Could not determine tax reference year from {sheet_name}: {label!r}"
        )
    return int(match.group(1))


def read_formula_evidence(
    workbook_path: Path,
    *,
    sheet_name: str,
    cells_by_code: Mapping[str, str],
) -> dict[str, str]:
    """Read targeted source formulas without changing cached-value extraction.

    Each requested cell is tied to the municipality code in column A on the
    same row. This makes a moved row or replaced formula fail with source-cell
    context instead of silently attaching evidence to the wrong municipality.
    """

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
        keep_vba=True,
        keep_links=True,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ExtractionError(f"Missing source sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        formulas: dict[str, str] = {}
        seen_cells: set[str] = set()

        for expected_code, source_cell in cells_by_code.items():
            code = canonical_code(expected_code)
            if code not in TARGET_MUNICIPALITIES:
                raise ExtractionError(
                    f"Formula evidence requested for unknown target code {code!r}."
                )
            normalized_cell = source_cell.upper()
            if normalized_cell in seen_cells:
                raise ExtractionError(
                    f"Formula evidence cell is duplicated: "
                    f"{sheet_name}!{normalized_cell}."
                )
            seen_cells.add(normalized_cell)

            cell = sheet[normalized_cell]
            row_code = canonical_code(sheet.cell(row=cell.row, column=1).value)
            if row_code != code:
                raise ExtractionError(
                    f"{sheet_name}!{normalized_cell} was expected to belong to "
                    f"municipality code {code}; found {row_code!r} in column A."
                )

            formula = cell.value
            if cell.data_type != "f" or not isinstance(formula, str):
                raise ExtractionError(
                    f"{sheet_name}!{normalized_cell} for municipality code "
                    f"{code} was expected to contain a formula; found "
                    f"{formula!r}."
                )
            formulas[code] = formula

        return formulas
    finally:
        workbook.close()


def extract_records(workbook_path: Path) -> list[dict[str, Any]]:
    """Return one raw record for every target municipality-year.

    The workbook is opened read-only and with ``data_only=True`` so extraction
    reads the values distributed by NJ DCA rather than executing macros or
    recalculating formulas.
    """

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
        keep_vba=True,
        keep_links=True,
    )
    records: list[dict[str, Any]] = []
    try:
        for year in TARGET_YEARS:
            sheet_name = f"{year} Summary"
            if sheet_name not in workbook.sheetnames:
                raise ExtractionError(f"Missing annual source sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            header_values = next(
                sheet.iter_rows(
                    min_row=HEADER_ROW,
                    max_row=HEADER_ROW,
                    values_only=True,
                )
            )
            headers = _header_map(header_values)
            population_label, population_column = _require_matching_header(
                headers,
                sheet_name=sheet_name,
                description="population",
                predicate=lambda label: (
                    "population" in label.casefold()
                    and "density" not in label.casefold()
                ),
            )
            tax_label, tax_column = _require_matching_header(
                headers,
                sheet_name=sheet_name,
                description="tax-collection percentage",
                predicate=lambda label: label.startswith(
                    "% of Taxes Collected, CY"
                ),
            )

            columns = {
                # Column A preserves the source's zero-padded code. Column B
                # repeats it numerically, so relying on the duplicate header
                # text would select the wrong representation.
                "municipality_code": 1,
                "municipality_name": _require_header(
                    headers, "Municipality", sheet_name
                ),
                "no_ufb_available": _require_header(
                    headers, "No UFB Available", sheet_name
                ),
                "significant_data_missing": _require_header(
                    headers, "Sig. Data Missing", sheet_name
                ),
                "population_estimate": population_column,
                "tax_collection_pct": tax_column,
                "net_debt": _require_header(headers, "Net Debt", sheet_name),
                "per_capita_net_debt": _require_header(
                    headers, "Per Capita Net Debt", sheet_name
                ),
                "three_year_average_property_valuation": _require_header(
                    headers, "3 Yr. Average Property Valuation", sheet_name
                ),
                "net_debt_to_valuation_pct": _require_header(
                    headers,
                    "Net Debt as % of 3 Year Avg Property Valuation",
                    sheet_name,
                ),
                "total_structural_imbalances": _require_header(
                    headers, "Total Imbalances", sheet_name
                ),
            }
            max_column = max(columns.values())
            found: set[str] = set()

            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=FIRST_DATA_ROW,
                    max_col=max_column,
                    values_only=True,
                ),
                start=FIRST_DATA_ROW,
            ):
                code = canonical_code(row[columns["municipality_code"] - 1])
                if code not in TARGET_MUNICIPALITIES:
                    continue
                if code in found:
                    raise ExtractionError(
                        f"Duplicate target code {code} in {sheet_name}."
                    )
                found.add(code)
                record: dict[str, Any] = {
                    "municipality_code": code,
                    "source_municipality_name": row[
                        columns["municipality_name"] - 1
                    ],
                    "budget_year": year,
                    "population_source_label": population_label,
                    "tax_collection_source_label": tax_label,
                    "tax_collection_reference_year": _tax_reference_year(
                        tax_label, sheet_name
                    ),
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                    "source_workbook": SOURCE_WORKBOOK_NAME,
                }
                for field, column in columns.items():
                    if field in {"municipality_code", "municipality_name"}:
                        continue
                    record[field] = row[column - 1]
                    record[f"{field}_source_cell"] = _cell_address(
                        column, row_number
                    )
                records.append(record)

            missing = set(TARGET_MUNICIPALITIES) - found
            if missing:
                raise ExtractionError(
                    f"{sheet_name} is missing target codes: {sorted(missing)}"
                )
    finally:
        workbook.close()
    return records
